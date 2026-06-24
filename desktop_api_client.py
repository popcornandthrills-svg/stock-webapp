import json
import re
from datetime import datetime
from pathlib import Path
from urllib import error, parse, request

from openpyxl import load_workbook

from stock_core import BRANCH_POINTS, POINTS, invoice_branch_prefix, normalize_key


def invoice_is_returned(row):
    status = str((row or {}).get("status") or "").strip().lower()
    return status == "returned"


def invoice_effective_amounts(row):
    total_amount = max(float((row or {}).get("total_amount") or 0), 0)
    paid_amount = max(float((row or {}).get("paid_amount") or 0), 0)
    returned_total = min(max(float((row or {}).get("returned_total_amount") or 0), 0), total_amount)
    effective_total = round(max(total_amount - returned_total, 0), 2)
    effective_paid = round(min(paid_amount, effective_total), 2)
    effective_due = round(max(effective_total - effective_paid, 0), 2)
    return effective_total, effective_paid, effective_due


class ApiDB:
    def __init__(self, base_url, timeout=45):
        self.base_url = str(base_url or "").strip().rstrip("/")
        self.timeout = int(timeout)
        self.token = ""
        self.user = {}
        self._point_map = {name: idx for idx, name in enumerate(POINTS, start=1)}
        self.clear_cache()

    def clear_cache(self):
        self._inventory_cache = None
        self._moves_cache = None
        self._invoice_cache = None
        self._invoice_detail_cache = {}
        self._shop_manager_cache = None
        self._audit_log_cache = None
        self._sales_cache = {}

    def point_map(self):
        return dict(self._point_map)

    def _branch_name(self, branch_id):
        if branch_id is None:
            return ""
        try:
            branch_id = int(branch_id)
        except Exception:
            return ""
        for name, value in self._point_map.items():
            if value == branch_id:
                return name
        return ""

    def _auth_headers(self):
        headers = {"Accept": "application/json"}
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        return headers

    def _request(self, method, path, params=None, payload=None):
        if not self.base_url:
            raise ValueError("API URL is not configured")
        query = {}
        for key, value in (params or {}).items():
            if value is None:
                continue
            if isinstance(value, bool):
                query[key] = "true" if value else "false"
            else:
                query[key] = str(value)
        url = f"{self.base_url}{path}"
        if query:
            url = f"{url}?{parse.urlencode(query)}"
        data = None
        headers = self._auth_headers()
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
            headers["Content-Type"] = "application/json"
        req = request.Request(url, data=data, headers=headers, method=method.upper())
        try:
            with request.urlopen(req, timeout=self.timeout) as resp:
                raw = resp.read().decode("utf-8").strip()
        except error.HTTPError as exc:
            detail = ""
            try:
                body = exc.read().decode("utf-8").strip()
                if body:
                    parsed = json.loads(body)
                    detail = parsed.get("detail") or body
            except Exception:
                detail = exc.reason or str(exc)
            raise ValueError(detail or f"API error {exc.code}") from exc
        except error.URLError as exc:
            reason = getattr(exc, "reason", None)
            raise ConnectionError(f"Unable to reach backend API: {reason or exc}") from exc
        if not raw:
            return {}
        return json.loads(raw)

    def login(self, mode, password, branch=None):
        payload = {"mode": str(mode or "").strip(), "password": str(password or "").strip()}
        if branch:
            payload["branch"] = str(branch).strip()
        result = self._request("POST", "/auth/login", payload=payload)
        self.clear_cache()
        self.token = str(result.get("access_token") or "")
        self.user = {
            "role": str(result.get("role") or ""),
            "user_name": str(result.get("user_name") or ""),
            "branch_name": str(result.get("branch_name") or ""),
            "branch_id": result.get("branch_id"),
        }
        return result

    def logout(self):
        try:
            if self.token:
                self._request("POST", "/auth/logout")
        finally:
            self.token = ""
            self.user = {}
            self.clear_cache()

    def add_audit_log(self, *args, **kwargs):
        return None

    def has_inventory_cache(self):
        return self._inventory_cache is not None

    def has_moves_cache(self):
        return self._moves_cache is not None

    def has_invoice_cache(self):
        return self._invoice_cache is not None

    def has_invoice_details_cache(self, invoice_id):
        try:
            invoice_id = int(invoice_id)
        except Exception:
            return False
        return invoice_id in self._invoice_detail_cache

    def has_shop_manager_cache(self):
        return self._shop_manager_cache is not None

    def has_audit_log_cache(self):
        return self._audit_log_cache is not None

    def has_sales_cache(self, point_id=None):
        return self._sales_cache.get(str(point_id or "")) is not None

    def _default_inventory_summary(self):
        return {"skus": 0, "units": 0, "wholesale": 0.0, "retail": 0.0, "low": []}

    def _default_sales_summary(self):
        return {"sales_units": 0, "sales_entries": 0, "sales_value": 0.0, "top_moving": [], "top_overstock": []}

    def _shape_inventory_rows(self, rows):
        items = []
        for row in rows:
            shaped = dict(row)
            by = {name: int((row.get("by") or {}).get(name, 0) or 0) for name in POINTS}
            shaped["by"] = by
            shaped["total"] = int(row.get("total") or sum(by.values()))
            items.append(shaped)
        return items

    def cached_inventory(self):
        return [dict(row, by=dict(row.get("by") or {})) for row in (self._inventory_cache or [])]

    def inventory(self, force=False):
        if not self.token:
            return []
        if self._inventory_cache is None or force:
            result = self._request("GET", "/inventory", params={"limit": 1000})
            self._inventory_cache = self._shape_inventory_rows(result.get("items", []))
        return self.cached_inventory()

    def analytics(self, force=False):
        inv = self.inventory(force=force)
        if not inv:
            return self._default_inventory_summary()
        return {
            "skus": len(inv),
            "units": sum(int(row.get("total") or 0) for row in inv),
            "wholesale": round(sum(int(row.get("total") or 0) * float(row.get("wholesale") or 0) for row in inv), 2),
            "retail": round(sum(int(row.get("total") or 0) * float(row.get("retail") or 0) for row in inv), 2),
            "low": [row for row in inv if int(row.get("total") or 0) <= int(row.get("reorder_level") or 0)],
        }

    def sales_insights(self, point_id=None, force=False):
        if not self.token:
            return self._default_sales_summary()
        cache_key = str(point_id or "")
        if force or cache_key not in self._sales_cache:
            params = {}
            branch_name = self._branch_name(point_id)
            if branch_name:
                params["branch"] = branch_name
            result = self._request("GET", "/analytics", params=params)
            self._sales_cache[cache_key] = result.get("sales") or self._default_sales_summary()
        return dict(self._sales_cache.get(cache_key) or self._default_sales_summary())

    def cached_recent_invoices(self, limit=25, branch_id=None, due_only=False):
        rows = [dict(row) for row in (self._invoice_cache or [])]
        branch_name = self._branch_name(branch_id)
        if branch_name:
            rows = [row for row in rows if str(row.get("branch_name") or "") == branch_name]
        if due_only:
            rows = [
                row
                for row in rows
                if invoice_effective_amounts(row)[2] > 0.009
            ]
        return rows[: int(limit)]

    def recent_invoices(self, limit=25, branch_id=None, due_only=False, force=False):
        if not self.token:
            return []
        if self._invoice_cache is None or force:
            result = self._request("GET", "/invoices", params={"limit": 400})
            self._invoice_cache = list(result.get("invoices", []))
        return self.cached_recent_invoices(limit=limit, branch_id=branch_id, due_only=due_only)

    def invoice_details(self, invoice_id, force=False):
        invoice_id = int(invoice_id)
        if force or invoice_id not in self._invoice_detail_cache:
            result = self._request("GET", f"/invoices/{invoice_id}")
            self._invoice_detail_cache[invoice_id] = (
                result.get("invoice") or {},
                result.get("lines") or [],
            )
        head, lines = self._invoice_detail_cache.get(invoice_id) or ({}, [])
        return dict(head), [dict(line) for line in lines]

    def update_invoice_file(self, invoice_id, invoice_file):
        return None

    def return_invoice(self, invoice_id, return_no=None):
        payload = {"return_no": str(return_no or "").strip()} if return_no else None
        result = self._request("POST", f"/invoices/{int(invoice_id)}/return", payload=payload)
        self.clear_cache()
        return result.get("returned") or {}

    def return_invoice_items(self, invoice_id, items, return_no=None):
        result = self._request(
            "POST",
            f"/invoices/{int(invoice_id)}/return",
            payload={
                "items": [{"invoice_line_id": int(item.get("invoice_line_id") or 0), "qty": int(item.get("qty") or 0)} for item in items or []],
                "return_no": str(return_no or "").strip(),
            },
        )
        self.clear_cache()
        return result.get("returned") or {}

    def list_shop_managers(self, force=False):
        if not self.token:
            return []
        if self._shop_manager_cache is None or force:
            result = self._request("GET", "/shop-managers")
            self._shop_manager_cache = list(result.get("items", []))
        return [dict(row) for row in (self._shop_manager_cache or [])]

    def shop_manager_for_branch(self, branch_id):
        branch_id = int(branch_id)
        for row in self.list_shop_managers():
            if int(row.get("branch_id") or 0) == branch_id:
                return row
        return None

    def upsert_shop_manager(self, branch_id, manager_name, password):
        result = self._request(
            "POST",
            "/shop-managers",
            payload={
                "branch": self._branch_name(branch_id),
                "manager_name": str(manager_name or "").strip(),
                "password": str(password or "").strip(),
            },
        )
        self._shop_manager_cache = None
        self._audit_log_cache = None
        return result.get("item") or {}

    def reset_shop_manager_password(self, branch_id, password):
        result = self._request(
            "POST",
            f"/shop-managers/{parse.quote(self._branch_name(branch_id), safe='')}/reset-password",
            payload={"password": str(password or "").strip()},
        )
        self._shop_manager_cache = None
        self._audit_log_cache = None
        return result.get("item") or {}

    def delete_shop_manager(self, branch_id):
        result = self._request("DELETE", f"/shop-managers/{parse.quote(self._branch_name(branch_id), safe='')}")
        self._shop_manager_cache = None
        self._audit_log_cache = None
        return result.get("deleted") or {}

    def recent_audit_logs(self, limit=400, force=False):
        if not self.token:
            return []
        if self._audit_log_cache is None or force:
            result = self._request("GET", "/audit-logs", params={"limit": int(limit)})
            self._audit_log_cache = list(result.get("items", []))
        return [dict(row) for row in (self._audit_log_cache or [])][: int(limit)]

    def next_invoice_number(self, branch_name=None):
        branch_name = str(branch_name or "").strip() or ("H.O" if self.user.get("role") == "admin" else self.user.get("branch_name") or "")
        if not self.token:
            return f"{invoice_branch_prefix(branch_name)}-0001"
        prefix = f"{invoice_branch_prefix(branch_name)}-"
        rows = self.recent_invoices(limit=400)
        next_seq = 1
        for row in rows:
            invoice_no = str(row.get("invoice_no") or "")
            if not invoice_no.startswith(prefix):
                continue
            tail = invoice_no[len(prefix):]
            if tail.isdigit():
                next_seq = max(next_seq, int(tail) + 1)
        return f"{prefix}{next_seq:04d}"

    def billing_lookup_item(self, lookup, point_id):
        if not self.token:
            return None
        branch_name = self._branch_name(point_id)
        key = str(lookup or "").strip().upper()
        for row in self.inventory():
            if str(row.get("art_no") or "").upper() != key:
                continue
            return {
                "id": int(row.get("id") or 0),
                "art_no": str(row.get("art_no") or ""),
                "item_name": str(row.get("item_name") or ""),
                "category": str(row.get("category") or ""),
                "wholesale": float(row.get("wholesale") or 0),
                "retail": float(row.get("retail") or 0),
                "point_qty": int((row.get("by") or {}).get(branch_name, 0)),
            }
        return None

    def lookup_stock(self, lookup, point_id=1):
        if not self.token:
            return None
        branch_name = self._branch_name(point_id)
        key = str(lookup or "").strip().upper()
        for row in self.inventory():
            if str(row.get("art_no") or "").upper() != key:
                continue
            return {
                "id": int(row.get("id") or 0),
                "art_no": str(row.get("art_no") or ""),
                "item_name": str(row.get("item_name") or ""),
                "point_qty": int((row.get("by") or {}).get(branch_name, 0)),
            }
        return None

    def item_by_art(self, art_no):
        if not self.token:
            return None
        key = str(art_no or "").strip().upper()
        for row in self.inventory():
            if str(row.get("art_no") or "").upper() == key:
                return dict(row, by=dict(row.get("by") or {}))
        return None

    def delete_item(self, item_id):
        result = self._request("DELETE", f"/inventory/items/{int(item_id)}")
        self.clear_cache()
        return result.get("deleted") or {}

    def add_item(self, payload):
        result = self._request("POST", "/inventory/items", payload=dict(payload or {}))
        self.clear_cache()
        return str(result.get("action") or "")

    def transfer(self, lookup, from_id, to_id, qty, note):
        return self.multi_transfer(lookup, from_id, [(to_id, qty)], note)

    def multi_transfer(self, lookup, from_id, transfer_pairs, note):
        payload = {
            "lookup": str(lookup or "").strip(),
            "from_branch": self._branch_name(from_id),
            "note": str(note or "").strip(),
            "transfers": [{"to_branch": self._branch_name(to_id), "qty": int(qty)} for to_id, qty in transfer_pairs],
        }
        result = self._request("POST", "/stock/transfers", payload=payload)
        self.clear_cache()
        return int(result.get("moved_total") or 0)

    def moves(self, limit=150, branch_id=None, force=False):
        if not self.token:
            return []
        if self._moves_cache is None or force:
            result = self._request("GET", "/moves", params={"limit": 1000})
            self._moves_cache = list(result.get("moves", []))
        rows = [dict(row) for row in (self._moves_cache or [])]
        branch_name = self._branch_name(branch_id)
        if branch_name:
            rows = [row for row in rows if str(row.get("from_p") or "") == branch_name or str(row.get("to_p") or "") == branch_name]
        return rows[: int(limit)]

    def bulk_load_stock(self, rows, to_point_id, note, from_point_id=None):
        payload = {
            "branch": self._branch_name(to_point_id),
            "from_branch": self._branch_name(from_point_id) if from_point_id else None,
            "note": str(note or "").strip(),
            "rows": list(rows or []),
        }
        result = self._request("POST", "/stock/bulk-load", payload=payload)
        self.clear_cache()
        return int(result.get("processed") or 0), result.get("failed") or []

    def _sales_rows_from_excel(self, file_path):
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("Excel is empty")

        item_aliases = {"itemname", "name", "productname", "product", "item"}
        art_aliases = {"artno", "art", "artnumber", "articleno", "articlenumber", "article", "artcode", "itemcode"}
        price_aliases = {"price", "rate", "amount", "saleprice", "sellingprice", "retailprice", "wholesaleprice", "mrp"}
        qty_aliases = {"quantity", "qty", "qnty", "stockqty", "soldqty", "soldquantity", "pcs", "pieces"}

        item_col = art_col = price_col = qty_col = None
        data_start = 0
        for i, hdr in enumerate(all_rows[:25]):
            norm = [normalize_key(h) for h in hdr]
            item_col = next((j for j, key in enumerate(norm) if key in item_aliases), None)
            art_col = next((j for j, key in enumerate(norm) if key in art_aliases), None)
            price_col = next((j for j, key in enumerate(norm) if key in price_aliases), None)
            qty_col = next((j for j, key in enumerate(norm) if key in qty_aliases), None)
            if item_col is not None and art_col is not None and price_col is not None and qty_col is not None:
                data_start = i + 1
                break

        if item_col is None or art_col is None or price_col is None or qty_col is None:
            raise ValueError("Headers required: ITEM NAME, ART NO, PRICE, QUANTITY")

        rows = []
        for row_no, row in enumerate(all_rows[data_start:], start=data_start + 1):
            if row is None:
                continue
            rows.append(
                {
                    "row_no": row_no,
                    "art_no": str(row[art_col] if len(row) > art_col else "" or "").strip().upper(),
                    "item_name": str(row[item_col] if len(row) > item_col else "" or "").strip(),
                    "price": row[price_col] if len(row) > price_col else None,
                    "quantity": row[qty_col] if len(row) > qty_col else None,
                }
            )
        return rows

    def import_sales(self, point_id, file_path, sale_date):
        payload = {
            "branch": self._branch_name(point_id),
            "sale_date": str(sale_date or "").strip(),
            "source_name": f"Sales import {Path(file_path).name}",
            "rows": self._sales_rows_from_excel(file_path),
        }
        result = self._request("POST", "/sales/load", payload=payload)
        self.clear_cache()
        return int(result.get("processed") or 0), result.get("failed") or []

    def create_invoice(self, payload, lines):
        branch_name = str(payload.get("branch") or self._branch_name(payload.get("branch_id"))).strip()
        discount_mode = str(payload.get("discount_mode") or "Percent").strip() or "Percent"
        discount_value = payload.get("discount_value")
        if discount_value is None:
            discount_value = payload.get("discount_percent") if discount_mode == "Percent" else payload.get("discount_amount")
        tds_mode = str(payload.get("tds_mode") or "Percent").strip() or "Percent"
        tds_value = payload.get("tds_value")
        if tds_value is None:
            tds_value = payload.get("tds_percent") if tds_mode == "Percent" else payload.get("tds_amount")
        created_at = str(payload.get("created_at") or "").strip()
        date_value = str(payload.get("date") or "").strip()
        if not date_value and created_at:
            try:
                date_value = datetime.fromisoformat(created_at.replace("Z", "+00:00")).strftime("%d-%m-%Y")
            except Exception:
                date_value = ""
        request_payload = {
            "branch": branch_name,
            "customer_name": str(payload.get("customer_name") or "").strip() or "Walk-in Customer",
            "customer_phone": str(payload.get("customer_phone") or "").strip(),
            "address": str(payload.get("address") or "").strip(),
            "price_type": str(payload.get("price_type") or "Retail").strip() or "Retail",
            "date": date_value,
            "discount_mode": discount_mode,
            "discount_value": float(discount_value or 0),
            "tds_mode": tds_mode,
            "tds_value": float(tds_value or 0),
            "paid_amount": float(payload.get("paid_amount") or 0),
            "payment_mode": str(payload.get("payment_mode") or "Cash").strip() or "Cash",
            "note": str(payload.get("note") or "").strip(),
            "lines": [
                {
                    "item_id": int(line.get("item_id") or 0),
                    "qty": int(line.get("qty") or 0),
                    "unit": str(line.get("unit") or "Nos"),
                    "rate": float(line.get("rate") or 0),
                    "gst_percent": float(line.get("gst_percent") or 0),
                }
                for line in lines
            ],
        }
        result = self._request("POST", "/invoices", payload=request_payload)
        self.clear_cache()
        return result.get("saved") or {}
