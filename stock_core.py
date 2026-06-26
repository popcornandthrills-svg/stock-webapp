import sqlite3
import sys
import csv
import json
import os
import re
import calendar
from datetime import UTC, datetime, timedelta
from html import escape
from pathlib import Path

from openpyxl import load_workbook
try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:
    psycopg = None
    dict_row = None
try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

CATEGORIES = [
    "None",
    "Stone Item",
    "CZ Item",
    "Annamayya",
    "Temple Jewellery",
    "1st Quality Temple Jewellery",
    "TTC",
]

POINTS = ["H.O", "GPH", "GPC", "GPF", "MGP", "GPS", "AMARAVATHI", "KALIMATA", "NAVISHKA"]
BRANCH_POINTS = POINTS[1:] if len(POINTS) > 1 else POINTS
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
DB_PATH = APP_DIR / "stock.db"
INVOICE_DIR = APP_DIR / "invoices"
LOGIN_PASSWORD = "admin123"
BRAND_NAME = "GOLDPRINCE"
BRAND_COMPANY = "JEWELL INDUSTRY INDIA PVT. LTD."
BRAND_SINCE = "SINCE 1995"
BRAND_HEADER_LINE = f"{BRAND_COMPANY} | {BRAND_SINCE}"
BRAND_GOLD = "#c9a35d"
INVOICE_BRANCH_PREFIXES = {
    "H.O": "H.O",
    "AMARAVATHI": "AMT",
    "KALIMATA": "KT",
}
POSTGRES_PREFIXES = ("postgres://", "postgresql://")


def configure_tk_environment():
    if getattr(sys, "frozen", False):
        base_dir = Path(getattr(sys, "_MEIPASS", APP_DIR))
        tcl_dir = base_dir / "_tcl_data"
        tk_dir = base_dir / "_tk_data"
    else:
        base_dir = Path(sys.executable).resolve().parent
        tcl_dir = base_dir / "tcl" / "tcl8.6"
        tk_dir = base_dir / "tcl" / "tk8.6"
    if tcl_dir.exists():
        os.environ["TCL_LIBRARY"] = str(tcl_dir)
    if tk_dir.exists():
        os.environ["TK_LIBRARY"] = str(tk_dir)


def now_iso():
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def parse_scan_code(value):
    raw = str(value or "").strip().upper()
    parts = [p.strip() for p in raw.split("-")]
    if len(parts) != 3 or any(not p for p in parts):
        raise ValueError("Scan format must be: ARTNO-BATCHNO-DESIGNNO (example: A005-69957-99988)")
    return parts[0], parts[1], parts[2]


def decode_price(token):
    txt = str(token or "").strip()
    if len(txt) < 3:
        raise ValueError("Batch/Design must have at least 3 characters for price decoding")
    middle = txt[1:-1]
    digits = "".join(ch for ch in middle if ch.isdigit())
    if not digits:
        raise ValueError("Batch/Design middle part must contain digits for price decoding")
    return int(digits[::-1])


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def money(value):
    return f"Rs {float(value or 0):.2f}"


def effective_due_amount(total_amount, paid_amount):
    return round(max(float(total_amount or 0) - float(paid_amount or 0), 0), 2)


def invoice_is_returned(row):
    status = str((row or {}).get("status") or "").strip().lower()
    return status == "returned"


def invoice_returned_total(row):
    return round(max(float((row or {}).get("returned_total_amount") or 0), 0), 2)


def invoice_effective_amounts(row):
    total_amount = max(float((row or {}).get("total_amount") or 0), 0)
    paid_amount = max(float((row or {}).get("paid_amount") or 0), 0)
    returned_total = min(invoice_returned_total(row), total_amount)
    effective_total = round(max(total_amount - returned_total, 0), 2)
    effective_paid = round(min(paid_amount, effective_total), 2)
    effective_due = round(max(effective_total - effective_paid, 0), 2)
    return effective_total, effective_paid, effective_due


def next_return_number():
    return f"RET-{datetime.now().strftime('%Y%m%d-%H%M%S')}"


def display_time(value):
    txt = str(value or "").strip()
    if not txt:
        return ""
    try:
        dt = datetime.fromisoformat(txt.replace("Z", "+00:00"))
        return dt.astimezone().strftime("%d-%m-%Y %I:%M %p")
    except Exception:
        return txt.replace("T", " ")[:16]


def invoice_branch_prefix(branch_name):
    raw = str(branch_name or "").strip().upper()
    if not raw:
        return "INV"
    if raw in INVOICE_BRANCH_PREFIXES:
        return INVOICE_BRANCH_PREFIXES[raw]
    normalized = re.sub(r"[^A-Z0-9]+", "", raw)
    return normalized or "INV"


def is_postgres_target(value):
    return str(value or "").strip().lower().startswith(POSTGRES_PREFIXES)


class DBSession:
    def __init__(self, owner, conn):
        self.owner = owner
        self.conn = conn

    def execute(self, sql, params=None):
        query = self.owner._adapt_sql(sql)
        if params is None:
            return self.conn.execute(query)
        return self.conn.execute(query, params)

    def __enter__(self):
        self.conn.__enter__()
        return self

    def __exit__(self, exc_type, exc, tb):
        return self.conn.__exit__(exc_type, exc, tb)

    def __getattr__(self, name):
        return getattr(self.conn, name)


class DB:
    def __init__(self, path):
        self.is_postgres = is_postgres_target(path)
        if self.is_postgres:
            if psycopg is None:
                raise RuntimeError("psycopg is required for PostgreSQL/Neon connections")
            self.path = str(path).strip()
        else:
            self.path = Path(path).expanduser().resolve()
        self.init()

    def _adapt_sql(self, sql):
        query = str(sql)
        return query.replace("?", "%s") if self.is_postgres else query

    def _insert_and_get_id(self, db, sql, params):
        if self.is_postgres:
            row = db.execute(f"{str(sql).rstrip().rstrip(';')} RETURNING id", params).fetchone()
            return int(row["id"])
        cur = db.execute(sql, params)
        return int(cur.lastrowid)

    def c(self):
        if self.is_postgres:
            conn = psycopg.connect(self.path, row_factory=dict_row)
            return DBSession(self, conn)
        conn = sqlite3.connect(self.path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA busy_timeout = 30000")
        conn.execute("PRAGMA journal_mode = WAL")
        return DBSession(self, conn)

    def init(self):
        if self.is_postgres:
            return self._init_postgres()
        return self._init_sqlite()

    def _init_sqlite(self):
        with self.c() as db:
            db.execute("CREATE TABLE IF NOT EXISTS points(id INTEGER PRIMARY KEY, name TEXT UNIQUE, is_ho INTEGER DEFAULT 0)")
            db.execute(
                """CREATE TABLE IF NOT EXISTS items(
                    id INTEGER PRIMARY KEY,
                    art_no TEXT UNIQUE,
                    batch_no TEXT NOT NULL DEFAULT '',
                    design_no TEXT NOT NULL DEFAULT '',
                    item_name TEXT,
                    category TEXT,
                    wholesale REAL DEFAULT 0,
                    retail REAL DEFAULT 0,
                    reorder_level INTEGER DEFAULT 0,
                    description TEXT DEFAULT '',
                    created_at TEXT,
                    updated_at TEXT
                )"""
            )
            db.execute(
                "CREATE TABLE IF NOT EXISTS balances(point_id INTEGER,item_id INTEGER,qty INTEGER DEFAULT 0,PRIMARY KEY(point_id,item_id))"
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS moves(
                    id INTEGER PRIMARY KEY,
                    item_id INTEGER,
                    from_id INTEGER,
                    to_id INTEGER,
                    mtype TEXT,
                    qty INTEGER,
                    note TEXT,
                    created_at TEXT
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS invoices(
                    id INTEGER PRIMARY KEY,
                    invoice_no TEXT UNIQUE,
                    branch_id INTEGER,
                    customer_name TEXT DEFAULT '',
                    customer_phone TEXT DEFAULT '',
                    address TEXT DEFAULT '',
                    price_type TEXT DEFAULT 'Retail',
                    taxable_amount REAL DEFAULT 0,
                    cgst_amount REAL DEFAULT 0,
                    sgst_amount REAL DEFAULT 0,
                    subtotal REAL DEFAULT 0,
                    discount_percent REAL DEFAULT 0,
                    discount_amount REAL DEFAULT 0,
                    tds_percent REAL DEFAULT 0,
                    tds_amount REAL DEFAULT 0,
                    round_off REAL DEFAULT 0,
                    total_amount REAL DEFAULT 0,
                    paid_amount REAL DEFAULT 0,
                    due_amount REAL DEFAULT 0,
                    status TEXT DEFAULT 'Active',
                    returned_at TEXT DEFAULT '',
                    return_note TEXT DEFAULT '',
                    returned_total_amount REAL DEFAULT 0,
                    payment_mode TEXT DEFAULT 'Cash',
                    note TEXT DEFAULT '',
                    invoice_file TEXT DEFAULT '',
                    created_at TEXT
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS invoice_lines(
                    id INTEGER PRIMARY KEY,
                    invoice_id INTEGER NOT NULL,
                    item_id INTEGER,
                    art_no TEXT,
                    item_name TEXT,
                    unit TEXT DEFAULT 'Nos',
                    qty INTEGER DEFAULT 0,
                    rate REAL DEFAULT 0,
                    gst_percent REAL DEFAULT 0,
                    taxable_amount REAL DEFAULT 0,
                    line_total REAL DEFAULT 0,
                    FOREIGN KEY(invoice_id) REFERENCES invoices(id) ON DELETE CASCADE
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS shop_managers(
                    id INTEGER PRIMARY KEY,
                    branch_id INTEGER NOT NULL UNIQUE,
                    manager_name TEXT NOT NULL DEFAULT '',
                    password TEXT NOT NULL DEFAULT '',
                    created_at TEXT,
                    updated_at TEXT,
                    FOREIGN KEY(branch_id) REFERENCES points(id) ON DELETE CASCADE
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS audit_logs(
                    id INTEGER PRIMARY KEY,
                    event_type TEXT NOT NULL DEFAULT '',
                    role TEXT NOT NULL DEFAULT '',
                    actor_name TEXT NOT NULL DEFAULT '',
                    branch_id INTEGER,
                    status TEXT NOT NULL DEFAULT '',
                    note TEXT NOT NULL DEFAULT '',
                    created_at TEXT,
                    FOREIGN KEY(branch_id) REFERENCES points(id) ON DELETE SET NULL
                )"""
            )

            cols = {r["name"] for r in db.execute("PRAGMA table_info(items)").fetchall()}
            if "batch_no" not in cols:
                db.execute("ALTER TABLE items ADD COLUMN batch_no TEXT NOT NULL DEFAULT ''")
            if "design_no" not in cols:
                db.execute("ALTER TABLE items ADD COLUMN design_no TEXT NOT NULL DEFAULT ''")

            invoice_cols = {r["name"] for r in db.execute("PRAGMA table_info(invoices)").fetchall()}
            for col_name, col_def in [
                ("address", "TEXT DEFAULT ''"),
                ("taxable_amount", "REAL DEFAULT 0"),
                ("cgst_amount", "REAL DEFAULT 0"),
                ("sgst_amount", "REAL DEFAULT 0"),
                ("discount_percent", "REAL DEFAULT 0"),
                ("tds_percent", "REAL DEFAULT 0"),
                ("tds_amount", "REAL DEFAULT 0"),
                ("round_off", "REAL DEFAULT 0"),
                ("due_amount", "REAL DEFAULT 0"),
                ("status", "TEXT DEFAULT 'Active'"),
                ("returned_at", "TEXT DEFAULT ''"),
                ("return_note", "TEXT DEFAULT ''"),
                ("returned_total_amount", "REAL DEFAULT 0"),
                ("payment_mode", "TEXT DEFAULT 'Cash'"),
            ]:
                if col_name not in invoice_cols:
                    db.execute(f"ALTER TABLE invoices ADD COLUMN {col_name} {col_def}")

            line_cols = {r["name"] for r in db.execute("PRAGMA table_info(invoice_lines)").fetchall()}
            for col_name, col_def in [
                ("unit", "TEXT DEFAULT 'Nos'"),
                ("gst_percent", "REAL DEFAULT 0"),
                ("taxable_amount", "REAL DEFAULT 0"),
                ("returned_qty", "INTEGER DEFAULT 0"),
            ]:
                if col_name not in line_cols:
                    db.execute(f"ALTER TABLE invoice_lines ADD COLUMN {col_name} {col_def}")

            for i, name in enumerate(POINTS, start=1):
                db.execute(
                    "INSERT INTO points(id,name,is_ho) VALUES(?,?,?) ON CONFLICT(id) DO UPDATE SET name=excluded.name,is_ho=excluded.is_ho",
                    (i, name, 1 if name == "H.O" else 0),
                )

    def _init_postgres(self):
        with self.c() as db:
            db.execute("CREATE TABLE IF NOT EXISTS points(id INTEGER PRIMARY KEY, name TEXT UNIQUE, is_ho INTEGER DEFAULT 0)")
            db.execute(
                """CREATE TABLE IF NOT EXISTS items(
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    art_no TEXT UNIQUE,
                    batch_no TEXT NOT NULL DEFAULT '',
                    design_no TEXT NOT NULL DEFAULT '',
                    item_name TEXT,
                    category TEXT,
                    wholesale DOUBLE PRECISION DEFAULT 0,
                    retail DOUBLE PRECISION DEFAULT 0,
                    reorder_level INTEGER DEFAULT 0,
                    description TEXT DEFAULT '',
                    created_at TEXT,
                    updated_at TEXT
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS balances(
                    point_id INTEGER,
                    item_id INTEGER,
                    qty INTEGER DEFAULT 0,
                    PRIMARY KEY(point_id,item_id)
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS moves(
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    item_id INTEGER,
                    from_id INTEGER,
                    to_id INTEGER,
                    mtype TEXT,
                    qty INTEGER,
                    note TEXT,
                    created_at TEXT
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS invoices(
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    invoice_no TEXT UNIQUE,
                    branch_id INTEGER,
                    customer_name TEXT DEFAULT '',
                    customer_phone TEXT DEFAULT '',
                    address TEXT DEFAULT '',
                    price_type TEXT DEFAULT 'Retail',
                    taxable_amount DOUBLE PRECISION DEFAULT 0,
                    cgst_amount DOUBLE PRECISION DEFAULT 0,
                    sgst_amount DOUBLE PRECISION DEFAULT 0,
                    subtotal DOUBLE PRECISION DEFAULT 0,
                    discount_percent DOUBLE PRECISION DEFAULT 0,
                    discount_amount DOUBLE PRECISION DEFAULT 0,
                    tds_percent DOUBLE PRECISION DEFAULT 0,
                    tds_amount DOUBLE PRECISION DEFAULT 0,
                    round_off DOUBLE PRECISION DEFAULT 0,
                    total_amount DOUBLE PRECISION DEFAULT 0,
                    paid_amount DOUBLE PRECISION DEFAULT 0,
                    due_amount DOUBLE PRECISION DEFAULT 0,
                    payment_mode TEXT DEFAULT 'Cash',
                    note TEXT DEFAULT '',
                    invoice_file TEXT DEFAULT '',
                    created_at TEXT
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS invoice_lines(
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    invoice_id INTEGER NOT NULL,
                    item_id INTEGER,
                    art_no TEXT,
                    item_name TEXT,
                    unit TEXT DEFAULT 'Nos',
                    qty INTEGER DEFAULT 0,
                    rate DOUBLE PRECISION DEFAULT 0,
                    gst_percent DOUBLE PRECISION DEFAULT 0,
                    taxable_amount DOUBLE PRECISION DEFAULT 0,
                    line_total DOUBLE PRECISION DEFAULT 0,
                    FOREIGN KEY(invoice_id) REFERENCES invoices(id) ON DELETE CASCADE
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS shop_managers(
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    branch_id INTEGER NOT NULL UNIQUE,
                    manager_name TEXT NOT NULL DEFAULT '',
                    password TEXT NOT NULL DEFAULT '',
                    created_at TEXT,
                    updated_at TEXT,
                    FOREIGN KEY(branch_id) REFERENCES points(id) ON DELETE CASCADE
                )"""
            )
            db.execute(
                """CREATE TABLE IF NOT EXISTS audit_logs(
                    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    event_type TEXT NOT NULL DEFAULT '',
                    role TEXT NOT NULL DEFAULT '',
                    actor_name TEXT NOT NULL DEFAULT '',
                    branch_id INTEGER,
                    status TEXT NOT NULL DEFAULT '',
                    note TEXT NOT NULL DEFAULT '',
                    created_at TEXT,
                    FOREIGN KEY(branch_id) REFERENCES points(id) ON DELETE SET NULL
                )"""
            )

            db.execute("ALTER TABLE items ADD COLUMN IF NOT EXISTS batch_no TEXT NOT NULL DEFAULT ''")
            db.execute("ALTER TABLE items ADD COLUMN IF NOT EXISTS design_no TEXT NOT NULL DEFAULT ''")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS address TEXT DEFAULT ''")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS taxable_amount DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS cgst_amount DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS sgst_amount DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS discount_percent DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS tds_percent DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS tds_amount DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS round_off DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS due_amount DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS payment_mode TEXT DEFAULT 'Cash'")
            db.execute("ALTER TABLE invoice_lines ADD COLUMN IF NOT EXISTS unit TEXT DEFAULT 'Nos'")
            db.execute("ALTER TABLE invoice_lines ADD COLUMN IF NOT EXISTS gst_percent DOUBLE PRECISION DEFAULT 0")
            db.execute("ALTER TABLE invoice_lines ADD COLUMN IF NOT EXISTS taxable_amount DOUBLE PRECISION DEFAULT 0")

            for i, name in enumerate(POINTS, start=1):
                db.execute(
                    "INSERT INTO points(id,name,is_ho) VALUES(?,?,?) ON CONFLICT(id) DO UPDATE SET name=excluded.name,is_ho=excluded.is_ho",
                    (i, name, 1 if name == "H.O" else 0),
                )

    def point_map(self):
        with self.c() as db:
            rows = db.execute("SELECT id,name FROM points ORDER BY id").fetchall()
        return {r["name"]: r["id"] for r in rows}

    def branch_rows(self):
        with self.c() as db:
            rows = db.execute("SELECT id,name FROM points WHERE is_ho=0 ORDER BY id").fetchall()
        return [dict(r) for r in rows]

    def shop_manager_for_branch(self, branch_id):
        with self.c() as db:
            row = db.execute(
                """
                SELECT sm.id, sm.branch_id, sm.manager_name, sm.password, sm.created_at, sm.updated_at, p.name branch_name
                FROM shop_managers sm
                JOIN points p ON p.id=sm.branch_id
                WHERE sm.branch_id=?
                """,
                (int(branch_id),),
            ).fetchone()
        return dict(row) if row else None

    def list_shop_managers(self):
        with self.c() as db:
            rows = db.execute(
                """
                SELECT sm.id, sm.branch_id, sm.manager_name, sm.created_at, sm.updated_at, p.name branch_name
                FROM shop_managers sm
                JOIN points p ON p.id=sm.branch_id
                ORDER BY p.id
                """
            ).fetchall()
        return [dict(r) for r in rows]

    def upsert_shop_manager(self, branch_id, manager_name, password):
        branch_id = int(branch_id)
        manager_name = str(manager_name or "").strip()
        password = str(password or "").strip()
        if branch_id <= 1:
            raise ValueError("Select a valid branch for the shop manager")
        if not manager_name:
            raise ValueError("Shop manager name is required")
        if not password:
            raise ValueError("Password is required")

        ts = now_iso()
        with self.c() as db:
            branch = db.execute("SELECT id,name FROM points WHERE id=? AND is_ho=0", (branch_id,)).fetchone()
            if not branch:
                raise ValueError("Selected branch is not valid for shop manager login")
            existing = db.execute("SELECT id FROM shop_managers WHERE branch_id=?", (branch_id,)).fetchone()
            if existing:
                db.execute(
                    "UPDATE shop_managers SET manager_name=?, password=?, updated_at=? WHERE branch_id=?",
                    (manager_name, password, ts, branch_id),
                )
                action = "updated"
                manager_id = int(existing["id"])
            else:
                manager_id = self._insert_and_get_id(
                    db,
                    "INSERT INTO shop_managers(branch_id, manager_name, password, created_at, updated_at) VALUES(?,?,?,?,?)",
                    (branch_id, manager_name, password, ts, ts),
                )
                action = "created"
        return {
            "id": manager_id,
            "branch_id": branch_id,
            "branch_name": str(branch["name"]),
            "manager_name": manager_name,
            "action": action,
        }

    def reset_shop_manager_password(self, branch_id, password):
        branch_id = int(branch_id)
        password = str(password or "").strip()
        if branch_id <= 1:
            raise ValueError("Select a valid branch")
        if not password:
            raise ValueError("Enter the new password")

        ts = now_iso()
        with self.c() as db:
            current = db.execute(
                """
                SELECT sm.id, sm.manager_name, p.name branch_name
                FROM shop_managers sm
                JOIN points p ON p.id=sm.branch_id
                WHERE sm.branch_id=?
                """,
                (branch_id,),
            ).fetchone()
            if not current:
                raise ValueError("No shop manager login found for the selected branch")
            db.execute("UPDATE shop_managers SET password=?, updated_at=? WHERE branch_id=?", (password, ts, branch_id))
        return {
            "id": int(current["id"]),
            "branch_id": branch_id,
            "branch_name": str(current["branch_name"]),
            "manager_name": str(current["manager_name"] or ""),
        }

    def delete_shop_manager(self, branch_id):
        branch_id = int(branch_id)
        with self.c() as db:
            current = db.execute(
                """
                SELECT sm.id, sm.manager_name, p.name branch_name
                FROM shop_managers sm
                JOIN points p ON p.id=sm.branch_id
                WHERE sm.branch_id=?
                """,
                (branch_id,),
            ).fetchone()
            if not current:
                raise ValueError("No shop manager login found for the selected branch")
            db.execute("DELETE FROM shop_managers WHERE branch_id=?", (branch_id,))
        return {
            "id": int(current["id"]),
            "branch_id": branch_id,
            "branch_name": str(current["branch_name"]),
            "manager_name": str(current["manager_name"] or ""),
        }

    def verify_shop_manager(self, branch_id, password):
        branch_id = int(branch_id)
        password = str(password or "").strip()
        with self.c() as db:
            row = db.execute(
                """
                SELECT sm.id, sm.branch_id, sm.manager_name, p.name branch_name
                FROM shop_managers sm
                JOIN points p ON p.id=sm.branch_id
                WHERE sm.branch_id=? AND sm.password=?
                """,
                (branch_id, password),
            ).fetchone()
        return dict(row) if row else None

    def add_audit_log(self, event_type, role, actor_name="", branch_id=None, status="Success", note="", created_at=None):
        ts = str(created_at or now_iso())
        branch_value = int(branch_id) if branch_id is not None else None
        with self.c() as db:
            db.execute(
                """
                INSERT INTO audit_logs(event_type, role, actor_name, branch_id, status, note, created_at)
                VALUES(?,?,?,?,?,?,?)
                """,
                (
                    str(event_type or "").strip(),
                    str(role or "").strip(),
                    str(actor_name or "").strip(),
                    branch_value,
                    str(status or "").strip(),
                    str(note or "").strip(),
                    ts,
                ),
            )

    def recent_audit_logs(self, limit=400):
        with self.c() as db:
            rows = db.execute(
                """
                SELECT a.id, a.event_type, a.role, a.actor_name, a.status, a.note, a.created_at, p.name branch_name
                FROM audit_logs a
                LEFT JOIN points p ON p.id=a.branch_id
                ORDER BY a.id DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
        return [dict(r) for r in rows]

    def recent_audit_logs_for_art(self, art_no, limit=100):
        art = str(art_no or "").strip().upper()
        if not art:
            return []
        with self.c() as db:
            rows = db.execute(
                """
                SELECT a.id, a.event_type, a.role, a.actor_name, a.status, a.note, a.created_at, p.name branch_name
                FROM audit_logs a
                LEFT JOIN points p ON p.id=a.branch_id
                WHERE UPPER(a.note) LIKE ?
                ORDER BY a.id DESC
                LIMIT ?
                """,
                (f"%{art}%", int(limit)),
            ).fetchall()
        return [dict(r) for r in rows]

    def qty(self, db, point_id, item_id):
        row = db.execute(
            "SELECT qty FROM balances WHERE point_id=? AND item_id=?",
            (point_id, item_id),
        ).fetchone()
        return int(row["qty"]) if row else 0

    def adjust(self, db, point_id, item_id, delta):
        next_qty = self.qty(db, point_id, item_id) + int(delta)
        if next_qty < 0:
            raise ValueError("Insufficient stock")
        db.execute(
            "INSERT INTO balances(point_id,item_id,qty) VALUES(?,?,?) ON CONFLICT(point_id,item_id) DO UPDATE SET qty=excluded.qty",
            (point_id, item_id, next_qty),
        )

    def _find_by_art(self, db, art):
        rows = db.execute("SELECT * FROM items WHERE art_no=?", (art.upper(),)).fetchall()
        if len(rows) == 1:
            return rows[0]
        if len(rows) > 1:
            raise ValueError("Multiple items found for ART NO. Scan full code with batch and design.")
        return None

    def find_item(self, db, lookup):
        key = str(lookup or "").strip()
        if not key:
            return None
        if key.isdigit():
            row = db.execute("SELECT * FROM items WHERE id=?", (int(key),)).fetchone()
            if row:
                return row
        if "-" in key:
            try:
                art, batch, design = parse_scan_code(key)
                row = db.execute(
                    "SELECT * FROM items WHERE art_no=? AND batch_no=? AND design_no=?",
                    (art, batch, design),
                ).fetchone()
                if row:
                    return row
                return self._find_by_art(db, art)
            except ValueError:
                pass
        return self._find_by_art(db, key)

    def item_by_art(self, art_no):
        key = str(art_no or "").strip().upper()
        if not key:
            return None
        with self.c() as db:
            row = db.execute("SELECT * FROM items WHERE art_no=?", (key,)).fetchone()
        return dict(row) if row else None

    def delete_item(self, item_id):
        with self.c() as db:
            item = db.execute("SELECT id, art_no, item_name FROM items WHERE id=?", (int(item_id),)).fetchone()
            if not item:
                raise ValueError("Selected item not found")
            db.execute("DELETE FROM moves WHERE item_id=?", (int(item_id),))
            db.execute("DELETE FROM balances WHERE item_id=?", (int(item_id),))
            db.execute("DELETE FROM items WHERE id=?", (int(item_id),))
            return {
                "id": int(item["id"]),
                "art_no": str(item["art_no"] or ""),
                "item_name": str(item["item_name"] or ""),
            }

    def add_item(self, payload):
        art_no = str(payload.get("art_no", "")).strip().upper()
        batch_no = str(payload.get("batch_no", "")).strip().upper()
        design_no = str(payload.get("design_no", "")).strip().upper()
        item_name = str(payload.get("item_name", "")).strip()
        category = str(payload.get("category", "")).strip()
        reorder = int(payload.get("reorder", 0))
        quantity = int(payload.get("quantity", 0))
        branch_id = int(payload.get("branch_id") or 1)
        desc = str(payload.get("desc", "")).strip()

        if not art_no or not batch_no or not design_no:
            raise ValueError("ART NO, Batch No, and Design No are required")
        if quantity < 0:
            raise ValueError("Quantity cannot be negative")
        wholesale = decode_price(batch_no)
        retail = decode_price(design_no)

        ts = str(payload.get("created_at") or now_iso())
        with self.c() as db:
            existing = db.execute("SELECT * FROM items WHERE art_no=?", (art_no,)).fetchone()
            if existing:
                item_id = int(existing["id"])
                if not item_name:
                    item_name = str(existing["item_name"] or "").strip()
                if not item_name:
                    raise ValueError("Item Name is required")
                if category not in CATEGORIES:
                    category = str(existing["category"] or "").strip()
                if category not in CATEGORIES:
                    raise ValueError("Invalid category")
                if not desc:
                    desc = str(existing["description"] or "").strip()
                db.execute(
                    """
                    UPDATE items
                    SET batch_no=?, design_no=?, item_name=?, category=?, wholesale=?, retail=?, reorder_level=?, description=?, updated_at=?
                    WHERE id=?
                    """,
                    (
                        batch_no,
                        design_no,
                        item_name,
                        category,
                        float(wholesale),
                        float(retail),
                        reorder,
                        desc,
                        ts,
                        item_id,
                    ),
                )
                mode = "updated"
            else:
                if not item_name:
                    raise ValueError("Item Name is required")
                if category not in CATEGORIES:
                    raise ValueError("Invalid category")
                item_id = self._insert_and_get_id(
                    db,
                    """
                    INSERT INTO items(art_no,batch_no,design_no,item_name,category,wholesale,retail,reorder_level,description,created_at,updated_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (art_no, batch_no, design_no, item_name, category, float(wholesale), float(retail), reorder, desc, ts, ts),
                )
                mode = "created"
            if quantity > 0:
                self.adjust(db, branch_id, item_id, quantity)
                db.execute(
                    "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                    (
                        item_id,
                        None,
                        branch_id,
                        "ho_in" if branch_id == 1 else "branch_load",
                        quantity,
                        "Stock added from inventory form",
                        ts,
                    ),
                )
            return mode

    def inward_ho(self, lookup, qty, note):
        qty = int(qty)
        if qty <= 0:
            raise ValueError("Quantity must be > 0")
        with self.c() as db:
            item = self.find_item(db, lookup)
            if not item:
                raise ValueError("Item not found")
            self.adjust(db, 1, int(item["id"]), qty)
            db.execute(
                "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                (int(item["id"]), None, 1, "ho_in", qty, str(note or "").strip(), now_iso()),
            )

    def transfer(self, lookup, from_id, to_id, qty, note):
        qty = int(qty)
        from_id = int(from_id)
        to_id = int(to_id)
        if qty <= 0:
            raise ValueError("Quantity must be > 0")
        if from_id == to_id:
            raise ValueError("From and To cannot be same")
        with self.c() as db:
            item = self.find_item(db, lookup)
            if not item:
                raise ValueError("Item not found")
            self.adjust(db, from_id, int(item["id"]), -qty)
            self.adjust(db, to_id, int(item["id"]), qty)
            db.execute(
                "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                (int(item["id"]), from_id, to_id, "transfer", qty, str(note or "").strip(), now_iso()),
            )

    def multi_transfer(self, lookup, from_id, transfer_pairs, note):
        from_id = int(from_id)
        if not transfer_pairs:
            raise ValueError("Add at least one destination transfer")
        merged = {}
        for to_id, qty in transfer_pairs:
            to_id = int(to_id)
            qty = int(qty)
            if qty <= 0:
                raise ValueError("Each transfer quantity must be > 0")
            if to_id == from_id:
                raise ValueError("From and To cannot be same")
            merged[to_id] = merged.get(to_id, 0) + qty

        total = sum(merged.values())
        with self.c() as db:
            item = self.find_item(db, lookup)
            if not item:
                raise ValueError("Item not found")
            item_id = int(item["id"])
            self.adjust(db, from_id, item_id, -total)
            for to_id, qty in merged.items():
                self.adjust(db, to_id, item_id, qty)
                db.execute(
                    "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                    (item_id, from_id, int(to_id), "transfer", int(qty), str(note or "").strip(), now_iso()),
                )
        return total

    def lookup_stock(self, lookup, point_id=1):
        with self.c() as db:
            item = self.find_item(db, lookup)
            if not item:
                return None
            q = self.qty(db, int(point_id), int(item["id"]))
            return {
                "id": int(item["id"]),
                "art_no": str(item["art_no"] or ""),
                "item_name": str(item["item_name"] or ""),
                "point_qty": int(q),
            }

    def next_invoice_number(self, branch_name=None):
        prefix = f"{invoice_branch_prefix(branch_name)}-"
        with self.c() as db:
            row = db.execute(
                "SELECT invoice_no FROM invoices WHERE invoice_no LIKE ? ORDER BY id DESC LIMIT 1",
                (f"{prefix}%",),
            ).fetchone()
        next_seq = 1
        if row and row["invoice_no"]:
            tail = str(row["invoice_no"])[len(prefix):]
            if tail.isdigit():
                next_seq = int(tail) + 1
        return f"{prefix}{next_seq:04d}"

    def billing_lookup_item(self, lookup, point_id):
        with self.c() as db:
            item = self.find_item(db, lookup)
            if not item:
                return None
            item_id = int(item["id"])
            point_qty = self.qty(db, int(point_id), item_id)
            total_qty = 0
            for i, _name in enumerate(POINTS, start=1):
                total_qty += self.qty(db, i, item_id)
            return {
                "id": item_id,
                "art_no": str(item["art_no"] or ""),
                "item_name": str(item["item_name"] or ""),
                "category": str(item["category"] or ""),
                "wholesale": float(item["wholesale"] or 0),
                "retail": float(item["retail"] or 0),
                "point_qty": int(point_qty),
                "available_qty": int(total_qty),
            }

    def create_invoice(self, payload, lines):
        branch_id = int(payload.get("branch_id") or 0)
        if branch_id <= 0:
            raise ValueError("Select a valid branch")
        if not lines:
            raise ValueError("Add at least one billing row")

        customer_name = str(payload.get("customer_name") or "").strip() or "Walk-in Customer"
        customer_phone = str(payload.get("customer_phone") or "").strip()
        address = str(payload.get("address") or "").strip()
        price_type = str(payload.get("price_type") or "Retail").strip() or "Retail"
        taxable_amount = round(float(payload.get("taxable_amount") or 0), 2)
        cgst_amount = round(float(payload.get("cgst_amount") or 0), 2)
        sgst_amount = round(float(payload.get("sgst_amount") or 0), 2)
        subtotal = round(float(payload.get("subtotal") or 0), 2)
        discount_percent = round(float(payload.get("discount_percent") or 0), 2)
        discount_amount = round(float(payload.get("discount_amount") or 0), 2)
        tds_percent = round(float(payload.get("tds_percent") or 0), 2)
        tds_amount = round(float(payload.get("tds_amount") or 0), 2)
        round_off = round(float(payload.get("round_off") or 0), 2)
        total_amount = round(float(payload.get("total_amount") or 0), 2)
        paid_amount = round(float(payload.get("paid_amount") or total_amount), 2)
        due_amount = effective_due_amount(total_amount, paid_amount)
        payment_mode = str(payload.get("payment_mode") or "Cash").strip() or "Cash"
        note = str(payload.get("note") or "").strip()
        ts = str(payload.get("created_at") or now_iso())

        with self.c() as db:
            branch_row = db.execute("SELECT name FROM points WHERE id=?", (branch_id,)).fetchone()
            branch_name = str(branch_row["name"] or "") if branch_row else ""
            invoice_prefix_key = str(payload.get("invoice_prefix_key") or branch_name).strip()
            invoice_no = str(payload.get("invoice_no") or self.next_invoice_number(invoice_prefix_key)).strip().upper()
            invoice_id = self._insert_and_get_id(
                db,
                """
                INSERT INTO invoices(
                    invoice_no, branch_id, customer_name, customer_phone, address, price_type,
                    taxable_amount, cgst_amount, sgst_amount, subtotal, discount_percent, discount_amount,
                    tds_percent, tds_amount, round_off, total_amount, paid_amount, due_amount,
                    payment_mode, note, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    invoice_no,
                    branch_id,
                    customer_name,
                    customer_phone,
                    address,
                    price_type,
                    taxable_amount,
                    cgst_amount,
                    sgst_amount,
                    subtotal,
                    discount_percent,
                    discount_amount,
                    tds_percent,
                    tds_amount,
                    round_off,
                    total_amount,
                    paid_amount,
                    due_amount,
                    payment_mode,
                    note,
                    ts,
                ),
            )

            for line in lines:
                item_id = int(line.get("item_id") or 0)
                qty = int(line.get("qty") or 0)
                rate = round(float(line.get("rate") or 0), 2)
                if item_id <= 0 or qty <= 0:
                    raise ValueError("Invoice rows must have valid item and quantity")
                item = db.execute("SELECT id, art_no, item_name FROM items WHERE id=?", (item_id,)).fetchone()
                if not item:
                    raise ValueError("Invoice item no longer exists")
                balance_rows = db.execute(
                    "SELECT point_id, qty FROM balances WHERE item_id=? ORDER BY CASE WHEN point_id=? THEN 0 ELSE 1 END, point_id",
                    (item_id, branch_id),
                ).fetchall()
                available = sum(int(row["qty"] or 0) for row in balance_rows)
                if available < qty:
                    raise ValueError(f"Insufficient stock for {item['art_no']} at selected branch")
                remaining_qty = qty
                for balance_row in balance_rows:
                    if remaining_qty <= 0:
                        break
                    point_id = int(balance_row["point_id"])
                    point_qty = int(balance_row["qty"] or 0)
                    if point_qty <= 0:
                        continue
                    use_qty = min(point_qty, remaining_qty)
                    self.adjust(db, point_id, item_id, -use_qty)
                    db.execute(
                        "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                        (item_id, point_id, None, "sale", use_qty, f"Invoice {invoice_no}", ts),
                    )
                    remaining_qty -= use_qty
                gst_percent = round(float(line.get("gst_percent") or 0), 2)
                taxable_line = round(float(line.get("taxable_amount") or (qty * rate)), 2)
                line_total = round(float(line.get("line_total") or 0), 2)
                db.execute(
                    """
                    INSERT INTO invoice_lines(invoice_id, item_id, art_no, item_name, unit, qty, rate, gst_percent, taxable_amount, line_total)
                    VALUES(?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        invoice_id,
                        item_id,
                        str(line.get("art_no") or item["art_no"] or ""),
                        str(line.get("item_name") or item["item_name"] or ""),
                        str(line.get("unit") or "Nos"),
                        qty,
                        rate,
                        gst_percent,
                        taxable_line,
                        line_total,
                    ),
                )

        return {
            "id": invoice_id,
            "invoice_no": invoice_no,
            "created_at": ts,
            "total_amount": total_amount,
        }

    def recent_invoices(self, limit=25, branch_id=None, due_only=False):
        where = []
        params = []
        if branch_id is not None:
            where.append("inv.branch_id=?")
            params.append(int(branch_id))
        if due_only:
            where.append(
                "(COALESCE(inv.total_amount, 0) - COALESCE(inv.returned_total_amount, 0) - COALESCE(inv.paid_amount, 0)) > 0.009"
            )
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""
        with self.c() as db:
            rows = db.execute(
                f"""
                SELECT inv.id, inv.invoice_no, inv.customer_name, inv.customer_phone, inv.address,
                       inv.total_amount, inv.discount_amount, inv.paid_amount, inv.due_amount,
                       inv.status, inv.returned_at, inv.return_note, inv.returned_total_amount,
                       inv.created_at, inv.invoice_file, inv.price_type, inv.payment_mode, p.name branch_name
                FROM invoices inv
                LEFT JOIN points p ON p.id=inv.branch_id
                {where_sql}
                ORDER BY inv.id DESC
                LIMIT ?
                """,
                [*params, int(limit)],
            ).fetchall()
        return [dict(r) for r in rows]

    def invoice_details(self, invoice_id):
        with self.c() as db:
            head = db.execute(
                """
                SELECT inv.*, p.name branch_name
                FROM invoices inv
                LEFT JOIN points p ON p.id=inv.branch_id
                WHERE inv.id=?
                """,
                (int(invoice_id),),
            ).fetchone()
            if not head:
                raise ValueError("Invoice not found")
            rows = db.execute(
                """
                SELECT id, item_id, art_no, item_name, unit, qty, COALESCE(returned_qty, 0) returned_qty,
                       rate, gst_percent, taxable_amount, line_total
                FROM invoice_lines
                WHERE invoice_id=?
                ORDER BY id
                """,
                (int(invoice_id),),
            ).fetchall()
        return dict(head), [dict(r) for r in rows]

    def update_invoice_file(self, invoice_id, invoice_file):
        with self.c() as db:
            db.execute("UPDATE invoices SET invoice_file=? WHERE id=?", (str(invoice_file or ""), int(invoice_id)))

    def return_invoice(self, invoice_id, return_note="", return_no=None):
        head, lines = self.invoice_details(invoice_id)
        items = []
        for line in lines:
            remaining_qty = max(int(line.get("qty") or 0) - int(line.get("returned_qty") or 0), 0)
            if remaining_qty > 0:
                items.append({"invoice_line_id": int(line["id"]), "qty": remaining_qty})
        if not items:
            raise ValueError("Invoice already returned")
        return self.return_invoice_items(invoice_id, items, return_note=return_note, return_no=return_no)

    def return_invoice_items(self, invoice_id, items, return_note="", return_no=None):
        ts = now_iso()
        invoice_id = int(invoice_id)
        clean_note = str(return_note or "").strip()
        return_no = str(return_no or next_return_number()).strip().upper()
        with self.c() as db:
            head = db.execute(
                """
                SELECT id, invoice_no, branch_id, note, status, returned_at, total_amount, paid_amount, returned_total_amount
                FROM invoices
                WHERE id=?
                """,
                (invoice_id,),
            ).fetchone()
            if not head:
                raise ValueError("Invoice not found")
            head_dict = dict(head)
            branch_id = int(head_dict.get("branch_id") or 0)
            if branch_id <= 0:
                raise ValueError("Invoice branch is invalid")
            line_rows = db.execute(
                """
                SELECT id, item_id, art_no, item_name, qty, COALESCE(returned_qty, 0) returned_qty, taxable_amount, line_total
                FROM invoice_lines
                WHERE invoice_id=?
                ORDER BY id
                """,
                (invoice_id,),
            ).fetchall()
            if not line_rows:
                raise ValueError("Invoice has no items to return")
            lines = {int(row["id"]): dict(row) for row in line_rows}

            invoice_no = str(head_dict.get("invoice_no") or "").strip() or f"#{invoice_id}"
            move_note = f"Return {return_no} | Invoice {invoice_no}"
            if clean_note:
                move_note = f"{move_note} | {clean_note}"
            returned_total_amount = 0.0
            returned_any = False
            for item in items or []:
                line_id = int(item.get("invoice_line_id") or 0)
                qty = int(item.get("qty") or 0)
                if line_id <= 0 or qty <= 0:
                    continue
                line = lines.get(line_id)
                if not line:
                    raise ValueError("Selected invoice item no longer exists")
                item_id = int(line["item_id"] or 0)
                total_qty = int(line["qty"] or 0)
                prev_returned_qty = int(line.get("returned_qty") or 0)
                remaining_qty = max(total_qty - prev_returned_qty, 0)
                if qty > remaining_qty:
                    raise ValueError(f"Return qty exceeds remaining qty for {line.get('art_no') or line.get('item_name') or 'item'}")
                if item_id <= 0 or total_qty <= 0:
                    continue
                self.adjust(db, branch_id, item_id, qty)
                db.execute(
                    "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                    (item_id, None, branch_id, "return", qty, move_note, ts),
                )
                db.execute(
                    "UPDATE invoice_lines SET returned_qty=COALESCE(returned_qty, 0)+? WHERE id=?",
                    (qty, line_id),
                )
                line_total = float(line.get("line_total") or 0)
                returned_total_amount += round((line_total * qty) / float(total_qty), 2)
                returned_any = True

            if not returned_any:
                raise ValueError("Select at least one item quantity to return")

            existing_note = str(head_dict.get("note") or "").strip()
            audit_note = f"Returned on {display_time(ts)} | {return_no}"
            if clean_note:
                audit_note = f"{audit_note} | {clean_note}"
            merged_note = f"{existing_note} | {audit_note}" if existing_note else audit_note
            fresh_lines = db.execute(
                "SELECT qty, COALESCE(returned_qty, 0) returned_qty FROM invoice_lines WHERE invoice_id=?",
                (invoice_id,),
            ).fetchall()
            fully_returned = all(int(row["returned_qty"] or 0) >= int(row["qty"] or 0) for row in fresh_lines)
            new_returned_total = round(
                min(float(head_dict.get("returned_total_amount") or 0) + returned_total_amount, float(head_dict.get("total_amount") or 0)),
                2,
            )
            new_status = "Returned" if fully_returned else "Partial Return"
            total_amount = float(head_dict.get("total_amount") or 0)
            paid_amount = float(head_dict.get("paid_amount") or 0)
            effective_total = round(max(total_amount - new_returned_total, 0), 2)
            effective_paid = round(min(paid_amount, effective_total), 2)
            effective_due = round(max(effective_total - effective_paid, 0), 2)
            db.execute(
                """
                UPDATE invoices
                SET status=?,
                    returned_at=?,
                    return_note=?,
                    due_amount=?,
                    returned_total_amount=?,
                    note=?
                WHERE id=?
                """,
                (new_status, ts, clean_note, effective_due, new_returned_total, merged_note, invoice_id),
            )
        return {
            "id": invoice_id,
            "status": new_status,
            "returned_at": ts,
            "return_no": return_no,
            "returned_total_amount": new_returned_total,
            "effective_total_amount": effective_total,
            "effective_paid_amount": effective_paid,
            "effective_due_amount": effective_due,
        }

    def bulk_load_stock(self, rows, to_point_id, note, from_point_id=None):
        to_point_id = int(to_point_id)
        if to_point_id < 1:
            raise ValueError("Invalid branch selected")
        from_point_id = int(from_point_id) if from_point_id else None
        if from_point_id is not None and from_point_id < 1:
            raise ValueError("Invalid source branch selected")
        if from_point_id is not None and from_point_id == to_point_id:
            raise ValueError("From and To branches cannot be the same")
        ts = now_iso()
        ok = 0
        bad = []
        with self.c() as db:
            for row_no, row in enumerate(rows, start=2):
                art = str(row.get("art_no", "")).strip().upper()
                name = str(row.get("item_name", "")).strip()
                try:
                    wholesale = float(row.get("wholesale", 0) or 0)
                    retail = float(row.get("retail", 0) or 0)
                    qty = int(float(row.get("quantity", 0) or 0))
                except Exception:
                    bad.append((row_no, art, "Invalid number format"))
                    continue
                if not art or not name or qty <= 0:
                    bad.append((row_no, art, "ART NO, Item Name, Quantity required"))
                    continue

                existing = db.execute("SELECT * FROM items WHERE art_no=?", (art,)).fetchone()
                if existing:
                    item_id = int(existing["id"])
                    category = str(existing["category"] or "").strip() or CATEGORIES[0]
                    db.execute(
                        """
                        UPDATE items
                        SET item_name=?, category=?, wholesale=?, retail=?, updated_at=?
                        WHERE id=?
                        """,
                        (name, category, float(wholesale), float(retail), ts, item_id),
                    )
                else:
                    item_id = self._insert_and_get_id(
                        db,
                        """
                        INSERT INTO items(art_no,batch_no,design_no,item_name,category,wholesale,retail,reorder_level,description,created_at,updated_at)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (art, "", "", name, CATEGORIES[0], float(wholesale), float(retail), 0, "", ts, ts),
                    )

                if from_point_id is not None:
                    available = self.qty(db, from_point_id, item_id)
                    if available < qty:
                        bad.append((row_no, art, "Insufficient stock in source branch"))
                        continue
                    self.adjust(db, from_point_id, item_id, -qty)
                    self.adjust(db, to_point_id, item_id, qty)
                    mtype = "branch_transfer"
                    move_from_id = from_point_id
                else:
                    self.adjust(db, to_point_id, item_id, qty)
                    mtype = "ho_in" if to_point_id == 1 else "branch_load"
                    move_from_id = None
                db.execute(
                    "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                    (item_id, move_from_id, to_point_id, mtype, qty, str(note or "").strip(), ts),
                )
                ok += 1

        return ok, bad[:100]

    def _sale_timestamp(self, sale_date):
        txt = str(sale_date or "").strip()
        try:
            d = datetime.strptime(txt, "%Y-%m-%d").date()
            return datetime(d.year, d.month, d.day, 12, 0, 0, tzinfo=UTC).isoformat().replace("+00:00", "Z")
        except ValueError:
            return now_iso()

    def import_sales_rows(self, point_id, rows, sale_date, source_name="Sales import"):
        point_id = int(point_id)
        ts = self._sale_timestamp(sale_date)
        ok = 0
        bad = []
        source_label = str(source_name or "Sales import").strip() or "Sales import"

        def parse_float(raw):
            txt = str(raw or "").strip()
            cleaned = re.sub(r"[^0-9.\-]", "", txt)
            if not cleaned or cleaned in {"-", ".", "-."}:
                raise ValueError
            return float(cleaned)

        def normalize_name(raw):
            return re.sub(r"\s+", " ", str(raw or "").strip().lower())

        with self.c() as db:
            for index, raw_row in enumerate(rows, start=1):
                row = dict(raw_row or {})
                row_no = int(row.get("row_no") or index)
                art = str(row.get("art_no") or "").strip().upper()
                item_name = str(row.get("item_name") or "").strip()
                try:
                    price = parse_float(row.get("price"))
                except Exception:
                    price = -1
                try:
                    qval = parse_float(row.get("quantity"))
                    if qval <= 0 or abs(qval - int(qval)) > 0.000001:
                        raise ValueError
                    need = int(qval)
                except Exception:
                    need = -1
                if not art or not item_name or need <= 0 or price < 0:
                    bad.append((row_no, art, "Invalid row (need Item Name, ART NO, Price, Qty)"))
                    continue

                item_rows = db.execute("SELECT * FROM items WHERE art_no=? ORDER BY id", (art,)).fetchall()
                if not item_rows:
                    bad.append((row_no, art, "Item not found"))
                    continue

                normalized_name = normalize_name(item_name)
                exact_name_rows = [it for it in item_rows if normalize_name(it["item_name"]) == normalized_name]
                partial_name_rows = [it for it in item_rows if normalized_name and normalized_name in normalize_name(it["item_name"])]
                candidate_rows = exact_name_rows or partial_name_rows or list(item_rows)

                price_rows = []
                for item in candidate_rows:
                    ws_price = float(item["wholesale"] or 0)
                    rt_price = float(item["retail"] or 0)
                    if abs(ws_price - price) < 0.01 or abs(rt_price - price) < 0.01:
                        price_rows.append(item)
                if not price_rows:
                    if len(candidate_rows) == 1:
                        price_rows = candidate_rows
                    else:
                        price_rows = candidate_rows

                remaining = need
                for item in price_rows:
                    if remaining <= 0:
                        break
                    available = self.qty(db, point_id, int(item["id"]))
                    if available <= 0:
                        continue
                    take = min(available, remaining)
                    self.adjust(db, point_id, int(item["id"]), -take)
                    db.execute(
                        "INSERT INTO moves(item_id,from_id,to_id,mtype,qty,note,created_at) VALUES(?,?,?,?,?,?,?)",
                        (int(item["id"]), point_id, None, "sale", int(take), source_label, ts),
                    )
                    remaining -= take

                if remaining > 0:
                    bad.append((row_no, art, "Insufficient stock"))
                else:
                    ok += 1

        return ok, bad[:50]

    def import_sales(self, point_id, file_path, sale_date):
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("Excel is empty")

        item_aliases = {
            "itemname",
            "name",
            "productname",
            "product",
            "item",
        }
        art_aliases = {
            "artno",
            "art",
            "artnumber",
            "articleno",
            "articlenumber",
            "article",
            "artcode",
            "itemcode",
        }
        price_aliases = {
            "price",
            "rate",
            "amount",
            "saleprice",
            "sellingprice",
            "retailprice",
            "wholesaleprice",
            "mrp",
        }
        qty_aliases = {
            "quantity",
            "qty",
            "qnty",
            "stockqty",
            "soldqty",
            "soldquantity",
            "pcs",
            "pieces",
        }

        item_col = None
        art_col = None
        price_col = None
        qty_col = None
        data_start = 0
        for i, hdr in enumerate(all_rows[:25]):
            norm = [normalize_key(h) for h in hdr]
            n = next((j for j, key in enumerate(norm) if key in item_aliases), None)
            a = next((j for j, key in enumerate(norm) if key in art_aliases), None)
            p = next((j for j, key in enumerate(norm) if key in price_aliases), None)
            q = next((j for j, key in enumerate(norm) if key in qty_aliases), None)
            if n is not None and a is not None and p is not None and q is not None:
                item_col = n
                art_col = a
                price_col = p
                qty_col = q
                data_start = i + 1
                break

        if item_col is None or art_col is None or price_col is None or qty_col is None:
            raise ValueError("Headers required: ITEM NAME, ART NO, PRICE, QUANTITY")

        rows = []
        for row_no, r in enumerate(all_rows[data_start:], start=data_start + 1):
            if r is None:
                continue
            rows.append(
                {
                    "row_no": row_no,
                    "art_no": str(r[art_col] if len(r) > art_col else "" or "").strip().upper(),
                    "item_name": str(r[item_col] if len(r) > item_col else "" or "").strip(),
                    "price": r[price_col] if len(r) > price_col else None,
                    "quantity": r[qty_col] if len(r) > qty_col else None,
                }
            )
        return self.import_sales_rows(point_id, rows, sale_date, source_name=f"Sales import {Path(file_path).name}")

    def inventory(self):
        with self.c() as db:
            items = db.execute("SELECT * FROM items ORDER BY id DESC").fetchall()
            out = []
            for it in items:
                row = dict(it)
                by_point = {}
                total = 0
                for i, name in enumerate(POINTS, start=1):
                    q = self.qty(db, i, int(it["id"]))
                    by_point[name] = q
                    total += q
                row["by"] = by_point
                row["total"] = total
                out.append(row)
            return out

    def analytics(self):
        inv = self.inventory()
        since = (datetime.now(UTC) - timedelta(days=30)).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        with self.c() as db:
            sold = {
                int(r["item_id"]): int(r["q"] or 0)
                for r in db.execute(
                    "SELECT item_id,SUM(qty) q FROM moves WHERE mtype='sale' AND created_at>=? GROUP BY item_id",
                    (since,),
                ).fetchall()
            }
        for i in inv:
            i["sold30"] = sold.get(int(i["id"]), 0)
        return {
            "skus": len(inv),
            "units": sum(i["total"] for i in inv),
            "wholesale": round(sum(i["total"] * float(i["wholesale"]) for i in inv), 2),
            "retail": round(sum(i["total"] * float(i["retail"]) for i in inv), 2),
            "low": [i for i in inv if i["total"] <= int(i["reorder_level"])],
        }

    def sales_insights(self, point_id=None):
        since = (datetime.now(UTC) - timedelta(days=30)).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        pid = int(point_id) if point_id is not None else None

        with self.c() as db:
            item_rows = db.execute("SELECT id,art_no,item_name,retail FROM items").fetchall()
            if pid is None:
                bal_rows = db.execute("SELECT item_id,SUM(qty) q FROM balances GROUP BY item_id").fetchall()
                sold_rows = db.execute(
                    "SELECT item_id,SUM(qty) q,COUNT(*) c FROM moves WHERE mtype='sale' GROUP BY item_id"
                ).fetchall()
                sold30_rows = db.execute(
                    "SELECT item_id,SUM(qty) q FROM moves WHERE mtype='sale' AND created_at>=? GROUP BY item_id",
                    (since,),
                ).fetchall()
            else:
                bal_rows = db.execute("SELECT item_id,qty q FROM balances WHERE point_id=?", (pid,)).fetchall()
                sold_rows = db.execute(
                    "SELECT item_id,SUM(qty) q,COUNT(*) c FROM moves WHERE mtype='sale' AND from_id=? GROUP BY item_id",
                    (pid,),
                ).fetchall()
                sold30_rows = db.execute(
                    "SELECT item_id,SUM(qty) q FROM moves WHERE mtype='sale' AND from_id=? AND created_at>=? GROUP BY item_id",
                    (pid, since),
                ).fetchall()

        items = {
            int(r["id"]): {
                "art_no": r["art_no"],
                "item_name": r["item_name"],
                "retail": float(r["retail"] or 0),
            }
            for r in item_rows
        }
        available_by = {int(r["item_id"]): int(r["q"] or 0) for r in bal_rows}
        sold_all = {int(r["item_id"]): int(r["q"] or 0) for r in sold_rows}
        sold30 = {int(r["item_id"]): int(r["q"] or 0) for r in sold30_rows}

        total_sales_units = sum(sold_all.values())
        total_sales_entries = sum(int(r["c"] or 0) for r in sold_rows)
        total_sales_value = 0.0
        for item_id, qty in sold_all.items():
            meta = items.get(item_id)
            if not meta or qty <= 0:
                continue
            total_sales_value += meta["retail"] * qty

        moving = []
        for item_id, qty in sold_all.items():
            meta = items.get(item_id)
            if not meta or qty <= 0:
                continue
            moving.append(
                {
                    "art_no": meta["art_no"],
                    "item_name": meta["item_name"],
                    "sold_qty": int(qty),
                    "available": int(available_by.get(item_id, 0)),
                }
            )
        moving.sort(key=lambda r: r["sold_qty"], reverse=True)
        top_moving_art_nos = {row["art_no"] for row in moving[:10]}

        over = []
        for item_id, meta in items.items():
            available = int(available_by.get(item_id, 0))
            sold_qty = int(sold_all.get(item_id, 0))
            sold30_qty = int(sold30.get(item_id, 0))
            if str(meta["art_no"] or "") in top_moving_art_nos:
                continue
            if available <= 0 and sold_qty <= 0:
                continue
            score = available - (sold_qty * 2)
            over.append(
                {
                    "art_no": meta["art_no"],
                    "item_name": meta["item_name"],
                    "available": available,
                    "sold": sold_qty,
                    "sold30": sold30_qty,
                    "score": score,
                }
            )
        over.sort(key=lambda r: (r["score"], r["available"]), reverse=True)

        return {
            "sales_units": int(total_sales_units),
            "sales_entries": int(total_sales_entries),
            "sales_value": round(total_sales_value, 2),
            "top_moving": moving[:10],
            "top_overstock": over[:10],
        }

    def moves(self, limit=150, branch_id=None, art_no=None):
        art = str(art_no or "").strip().upper()
        with self.c() as db:
            if branch_id is None:
                if art:
                    rows = db.execute(
                        """
                        SELECT m.id,m.mtype,m.qty,m.note,m.created_at,i.art_no,i.category,i.item_name,
                               p1.name from_p,p2.name to_p
                        FROM moves m
                        JOIN items i ON i.id=m.item_id
                        LEFT JOIN points p1 ON p1.id=m.from_id
                        LEFT JOIN points p2 ON p2.id=m.to_id
                        WHERE UPPER(i.art_no)=?
                        ORDER BY m.id DESC
                        LIMIT ?
                        """,
                        (art, int(limit)),
                    ).fetchall()
                else:
                    rows = db.execute(
                        """
                        SELECT m.id,m.mtype,m.qty,m.note,m.created_at,i.art_no,i.category,i.item_name,
                               p1.name from_p,p2.name to_p
                        FROM moves m
                        JOIN items i ON i.id=m.item_id
                        LEFT JOIN points p1 ON p1.id=m.from_id
                        LEFT JOIN points p2 ON p2.id=m.to_id
                        ORDER BY m.id DESC
                        LIMIT ?
                        """,
                        (int(limit),),
                    ).fetchall()
            else:
                if art:
                    rows = db.execute(
                        """
                        SELECT m.id,m.mtype,m.qty,m.note,m.created_at,i.art_no,i.category,i.item_name,
                               p1.name from_p,p2.name to_p
                        FROM moves m
                        JOIN items i ON i.id=m.item_id
                        LEFT JOIN points p1 ON p1.id=m.from_id
                        LEFT JOIN points p2 ON p2.id=m.to_id
                        WHERE (m.from_id=? OR m.to_id=?) AND UPPER(i.art_no)=?
                        ORDER BY m.id DESC
                        LIMIT ?
                        """,
                        (int(branch_id), int(branch_id), art, int(limit)),
                    ).fetchall()
                else:
                    rows = db.execute(
                        """
                        SELECT m.id,m.mtype,m.qty,m.note,m.created_at,i.art_no,i.category,i.item_name,
                               p1.name from_p,p2.name to_p
                        FROM moves m
                        JOIN items i ON i.id=m.item_id
                        LEFT JOIN points p1 ON p1.id=m.from_id
                        LEFT JOIN points p2 ON p2.id=m.to_id
                        WHERE (m.from_id=? OR m.to_id=?)
                        ORDER BY m.id DESC
                        LIMIT ?
                        """,
                        (int(branch_id), int(branch_id), int(limit)),
                    ).fetchall()
        return [dict(r) for r in rows]


